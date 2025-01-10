import os
from datetime import datetime

from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, Alignment, Border, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


class Excel:
    _instance = None
    _initialized = False

    def __new__(cls):
        """싱글턴 패턴을 구현."""
        if cls._instance is None:
            cls._instance = super().__new__(cls)
        return cls._instance

    def __init__(self):
        if not self._initialized:
            self._initialized = True

    def load_excel_file(self, excel_filename):
        """엑셀 파일을 열고, 없으면 새로 생성"""
        if os.path.exists(excel_filename):
            return load_workbook(excel_filename)
        else:
            return None

    def sort_sheets_by_date(self, wb):
        """엑셀 워크북 내 시트를 날짜 기준으로 정렬"""
        sheetnames = wb.sheetnames
        sheetnames.sort(key=lambda x: datetime.strptime(x, "%Y-%m-%d"), reverse=True)
        return sheetnames

    def find_latest_excel_file(self, directory):
        """output 폴더에서 가장 최신의 엑셀 파일을 찾습니다."""
        files = [f for f in os.listdir(directory) if f.endswith('.xlsx')]

        # 금일 날짜 제외
        today_date = datetime.now().strftime('%Y-%m-%d')
        files = [f for f in files if not f.startswith(today_date)]
        if not files:
            return None

        # 최신 파일 찾기
        latest_file = max(files, key=lambda x: os.path.getmtime(os.path.join(directory, x)))
        return os.path.join(directory, latest_file)

    def copy_old_sheets(self, new_wb, latest_file):
        """기존 엑셀 파일에서 최근 6개 시트를 복사하여 새로운 워크북에 추가"""
        if not latest_file:
            return []  # 기존 파일이 없으면 빈 리스트 반환

        old_wb = load_workbook(latest_file)
        sheet_names = sorted(old_wb.sheetnames, reverse=True)[:6]  # 최근 6개 시트만 가져옴

        copied_sheets = []
        for sheet_name in sheet_names:
            old_ws = old_wb[sheet_name]
            new_ws = new_wb.create_sheet(title=sheet_name)

            for row in old_ws.iter_rows():
                for cell in row:
                    new_ws[cell.coordinate].value = cell.value
                    # 셀 스타일도 복사합니다.
                    new_cell = new_ws[cell.coordinate]

                    # 스타일 속성 복사
                    new_cell.font = Font(name=cell.font.name, size=cell.font.size, bold=cell.font.bold,
                                         italic=cell.font.italic, color=cell.font.color)
                    new_cell.alignment = Alignment(horizontal=cell.alignment.horizontal,
                                                   vertical=cell.alignment.vertical)
                    new_cell.border = Border(left=cell.border.left, right=cell.border.right, top=cell.border.top,
                                             bottom=cell.border.bottom)
                    new_cell.fill = PatternFill(fill_type=cell.fill.fill_type, fgColor=cell.fill.fgColor)
                    new_cell.number_format = cell.number_format

            # 셀 병합 정보 복사
            for merged_range in old_ws.merged_cells.ranges:
                new_ws.merge_cells(str(merged_range))

            # 테이블 복사
            for table in old_ws.tables.values():
                new_table = Table(displayName=table.displayName, ref=table.ref)

                # 테이블 스타일 복사
                new_table.tableStyleInfo = TableStyleInfo(
                    name=table.tableStyleInfo.name,
                    showFirstColumn=table.tableStyleInfo.showFirstColumn,
                    showLastColumn=table.tableStyleInfo.showLastColumn,
                    showRowStripes=table.tableStyleInfo.showRowStripes,
                    showColumnStripes=table.tableStyleInfo.showColumnStripes
                )
                new_ws.add_table(new_table)

            copied_sheets.append(sheet_name)

        return copied_sheets

    def export(self, path, sheet_name, title, columns, data):
        """OCR 결과를 엑셀 파일로 저장하며, 기존 시트를 유지하고 새 데이터를 추가"""

        # 새로운 워크북 생성
        wb = Workbook()

        # 새로운 시트 추가 (금일 시트가 가장 먼저 오도록)
        ws = wb.active
        ws.title = sheet_name

        # A1 셀을 전체 너비만큼 병합하고 타이틀 적용
        last_column = get_column_letter(len(columns))  # 마지막 컬럼 계산
        ws.merge_cells(f"A1:{last_column}1")  # A1 ~ 마지막 컬럼 1행까지 병합
        ws["A1"] = title

        # 타이틀 스타일 적용
        ws["A1"].font = Font(size=16, bold=True)  # 글씨 크기 키우고 Bold 처리
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")  # 중앙 정렬
        ws.row_dimensions[1].height = 25  # 행 높이 키우기

        # 헤더 삽입
        ws.append([column.header for column in columns])

        # 데이터 삽입
        for row in data:
            row = row + [""] * (len(columns) - len(row))
            ws.append(row)  # 부족한 부분은 빈칸 채우기

            row_data = {column.key: row[col_idx] for col_idx, column in enumerate(columns)}
            for col_idx, column in enumerate(columns):
                column.style_handler(ws.cell(ws.max_row, col_idx + 1), row_data)

        # 표 스타일 적용
        last_row = ws.max_row
        table_ref = f"A2:{last_column}{last_row}"
        table = Table(displayName=f"Table_{sheet_name.replace('-', '_')}", ref=table_ref)

        style = TableStyleInfo(name="TableStyleLight8", showFirstColumn=False, showLastColumn=False,
                               showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)

        # 기존 시트 복사 (금일 시트 뒤에 추가)
        directory = os.path.dirname(path)
        latest_file = self.find_latest_excel_file(directory)
        copied_sheets = self.copy_old_sheets(wb, latest_file)

        # 새로 추가된 시트가 첫 번째가 되도록 정렬
        wb._sheets = [wb[sheet_name]] + [wb[name] for name in copied_sheets]

        # 파일 저장
        wb.save(path)

        print(f"✅ 엑셀 파일 저장 완료: {path}")
