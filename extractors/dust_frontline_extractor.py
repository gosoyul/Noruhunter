from datetime import datetime, timedelta

from openpyxl.styles import PatternFill

from circle_member import CircleMemberManager, CircleMember
from config import ConfigKeys, ConfigManager, DEFAULT_CONFIG
from excel import ExcelColumn
from utils import window
from . import Extractor

MIN_DAILY_POINT = 800


class DustFrontlineExtractor(Extractor):
    def __init__(self):
        manager = CircleMemberManager()
        ocr_headers = [("nickname", "닉네임"),
                       ("higher_point_title", "단일 전투 최고 점수 제목"), ("total_point_title", "누적 점수 제목"),
                       ("level", "레벨"), ("position", "직위"),
                       ("higher_point", "단일 전투 최고 점수"), ("total_point", "누적 점수")]
        excel_columns = [
            ExcelColumn(
                key="position",
                header="직위",
                value=lambda data: data["position"],
                style_handler=self._cell_style_handler
            ),
            ExcelColumn(
                key="join_date",
                header="가입일",
                value=lambda data: manager.get_by_nickname(data["nickname"], "join_date"),
                style_handler=self._cell_style_handler
            ),
            ExcelColumn(
                key="join_period",
                header="가입기간",
                value=lambda data: manager.get_by_nickname(data["nickname"], "join_period"),
                style_handler=self._cell_style_handler
            ),
            ExcelColumn(
                key="nickname",
                header="닉네임",
                value=lambda data: data["nickname"],
                style_handler=self._cell_style_handler
            ),
            ExcelColumn(
                key="arcalive_id",
                header="아카라이브 ID",
                value=lambda data: manager.get_by_nickname(data["nickname"], "arcalive_id"),
                style_handler=self._cell_style_handler
            ),
            ExcelColumn(
                key="uid",
                header="UID",
                value=lambda data: manager.get_by_nickname(data["nickname"], "uid"),
                style_handler=self._cell_style_handler
            ),
            ExcelColumn(
                key="level",
                header="레벨",
                value=lambda data: data["level"],
                style_handler=self._cell_style_handler
            ),
            ExcelColumn(
                key="higher_point",
                header="단일 전투 최고 점수",
                value=lambda data: data["higher_point"],
                style_handler=self._cell_style_handler
            ),
            ExcelColumn(
                key="total_point",
                header="누적 점수",
                value=lambda data: data["total_point"],
                style_handler=self._cell_style_handler
            ),
            ExcelColumn(
                key="missing_point",
                header="부족 점수",
                value=self.calculate_missing_point,
                style_handler=self._cell_style_handler
            ),
            ExcelColumn(
                key="remark",
                header="비고",
                value=lambda data: manager.get_by_nickname(data["nickname"], "remark"),
                style_handler=self._cell_style_handler
            )
        ]

        super().__init__(name="흙먼지전선 추출", ocr_headers=ocr_headers, excel_columns=excel_columns, folder_name="output_dust")

    def extract(self, parent=None, additional_scroll_repeat=-5):
        config = ConfigManager()
        dust_start_date = config.get(ConfigKeys.DUST_START_DATE)

        if dust_start_date == DEFAULT_CONFIG.get(ConfigKeys.DUST_START_DATE):
            window.show_message("알림", "흙먼지 전선의 시작일을 설정해주세요.")
            return

        super().extract(parent, additional_scroll_repeat)

    def _fix_extracted_text(self, extracted_text):
        # is_fix_needed = len(extracted_text) % len(self.ocr_headers)

        header_size = len(self.ocr_headers)
        max_size = (len(extracted_text) // header_size) * header_size
        for i in range(0, max_size, header_size):

            # 단일 전투 최고 점수 제목
            if "단일 전투" not in extracted_text[i + 1]:
                extracted_text[i] += extracted_text[i + 1]
                del extracted_text[i + 1]

            # 누적 점수 제목
            if "누적 점수" not in extracted_text[i + 2]:
                extracted_text[i + 1] += extracted_text[i + 2]
                del extracted_text[i + 3]

            # 레벨
            if not extracted_text[i + 3].startswith("Lv"):
                extracted_text[i + 2] += extracted_text[i + 3]
                del extracted_text[i + 3]

            # 직위
            if not extracted_text[i + 4] in ["서클장", "서클원", "부서클장"]:
                if CircleMemberManager().get_by_nickname(extracted_text[i]) is None:
                    extracted_text[i + 3] += extracted_text[i + 4]
                del extracted_text[i + 4]

            # 단일 전투 최고 점수
            if not extracted_text[i + 5].isdigit():
                number = self.extract_numbers(extracted_text[i + 5])
                if number == "":
                    extracted_text[i + 5] = 0
                else:
                    extracted_text[i + 5] = number

            # 누적 점수
            if not extracted_text[i + 6].isdigit():
                number = self.extract_numbers(extracted_text[i + 6])
                if number == "":
                    extracted_text[i + 5] = 0
                else:
                    extracted_text[i + 6] = number

        return extracted_text

    def calculate_missing_point(self, data):
        circle_member_manager = CircleMemberManager()
        join_date = circle_member_manager.get_by_nickname(data["nickname"], "join_date")

        result = self.calculate_total_point_goal(join_date) - int(data["total_point"])
        return result if result > 0 else 0

    def calculate_total_point_goal(self, join_date=None):
        config = ConfigManager()

        # 가입일을 기준으로 목표 누적 공헌도 계산
        today = datetime.now()

        # 5시 기준으로 요일을 변경
        if today.hour < 5:
            today -= timedelta(days=1)  # 5시 이전이면 전날로 설정

        dust_start_date = datetime.strptime(config.get(ConfigKeys.DUST_START_DATE), "%Y-%m-%d")
        join_date = dust_start_date if not join_date else datetime.strptime(join_date, '%Y-%m-%d')

        days_joined = (today - (join_date if join_date > dust_start_date else dust_start_date)).days

        # 목표 누적 공헌도 = 90 * 가입일 수
        return config.get(ConfigKeys.DUST_POINT_LIMIT) * days_joined

    def _cell_style_handler(self, cell, row_data):
        circle_member_manager = CircleMemberManager()
        member: CircleMember = circle_member_manager.get_by_nickname(row_data["nickname"])

        if member is None:
            cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")  # 배경색
            return

        config = ConfigManager()
        join_period = CircleMemberManager().get_by_nickname(row_data["nickname"], "join_period")
        if join_period is None:
            return

        daily_point_avg = row_data["total_point"] / join_period
        if daily_point_avg >= config.get(ConfigKeys.DUST_POINT_LIMIT):
            cell.fill = PatternFill(start_color="FFDFDF", end_color="FFDFDF", fill_type="solid")  # 배경색

    def get_dynamic_ratio(self, hwnd=None):
        hwnd = hwnd if hwnd else window.find_window(ConfigManager().get(ConfigKeys.WINDOW_TITLE))
        _, _, client_width, client_height, _, _ = window.get_client_area(hwnd=hwnd)

        # 윈도우 해상도를 기준으로 비율 계산
        aspect_ratio = client_width / client_height

        # 가로 비율 (일반 해상도 vs 울트라와이드)
        if aspect_ratio >= 21 / 9:  # 울트라와이드 (21:9 이상)
            left_ratio = 1 / 8
            right_ratio = left_ratio * 0.4
            top_ratio = 0.07  # 상단 비율
            bottom_ratio = top_ratio * 1.28
            profile_ratio = top_ratio * 0.7
        else:  # 일반 해상도 (16:9, 16:10)
            left_ratio = 1 / 6
            right_ratio = left_ratio * 0.3
            top_ratio = 0.09  # 상단 비율
            bottom_ratio = top_ratio * 1.35
            profile_ratio = top_ratio * 0.77

        print(f"🔍 감지된 화면 비율: {aspect_ratio:.2f}")

        return aspect_ratio, left_ratio, right_ratio, top_ratio, bottom_ratio, profile_ratio

    def _on_before_export(self, hwnd=None):
        pass

    def _on_after_export(self, hwnd=None):
        pass

    def _on_before_extract(self, hwnd=None):
        pass

    def _on_after_extract(self, hwnd=None):
        pass
