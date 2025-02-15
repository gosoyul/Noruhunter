from datetime import datetime, timedelta

from openpyxl.styles import PatternFill

from circle_member import CircleMemberManager, CircleMember
from config import ConfigKeys, ConfigManager
from excel import ExcelColumn
from utils import window
from . import Extractor

MAX_DAILY_CONTRIB = 90


class CircleMemberExtractor(Extractor):
    def __init__(self):
        manager = CircleMemberManager()
        ocr_headers = [("nickname", "닉네임"), ("position", "직위"),
                       ("weekly_contrib", "이번 주 공헌도"), ("total_contrib", "누적 공헌도"),
                       ("status", "상태"), ("level", "레벨")]
        excel_columns = [
            ExcelColumn(
                key="position",
                header="직위",
                value=lambda data: data["position"],
                style_handler=self._cell_style_handler
            ),
            ExcelColumn(
                key="join_date",
                header="가입일",
                value=lambda data: manager.get_by_nickname(data["nickname"], "join_date"),
                style_handler=self._cell_style_handler
            ),
            ExcelColumn(
                key="join_period",
                header="가입기간",
                value=lambda data: manager.get_by_nickname(data["nickname"], "join_period"),
                style_handler=self._cell_style_handler
            ),
            ExcelColumn(
                key="nickname",
                header="닉네임",
                value=lambda data: data["nickname"],
                style_handler=self._cell_style_handler
            ),
            ExcelColumn(
                key="arcalive_id",
                header="아카라이브 ID",
                value=lambda data: manager.get_by_nickname(data["nickname"], "arcalive_id"),
                style_handler=self._cell_style_handler
            ),
            ExcelColumn(
                key="uid",
                header="UID",
                value=lambda data: manager.get_by_nickname(data["nickname"], "uid"),
                style_handler=self._cell_style_handler
            ),
            ExcelColumn(
                key="level",
                header="레벨",
                value=lambda data: data["level"],
                style_handler=self._cell_style_handler
            ),
            ExcelColumn(
                key="weekly_contrib",
                header="이번 주 공헌도",
                value=lambda data: data["weekly_contrib"],
                style_handler=self._cell_style_handler
            ),
            ExcelColumn(
                key="total_contrib",
                header="누적 공헌도",
                value=lambda data: data["total_contrib"],
                style_handler=self._cell_style_handler
            ),
            ExcelColumn(
                key="missing_weekly_contrib",
                header="부족 공헌도",
                value=self.calculate_missing_weekly_contrib,
                style_handler=self._cell_style_handler
            ),
            ExcelColumn(
                key="status",
                header="상태",
                value=lambda data: data["status"],
                style_handler=self._cell_style_handler
            ),
            ExcelColumn(
                key="remark",
                header="비고",
                value=lambda data: manager.get_by_nickname(data["nickname"], "remark"),
                style_handler=self._cell_style_handler
            )
        ]

        super().__init__(name="서클원 추출", ocr_headers=ocr_headers, excel_columns=excel_columns)

    def _fix_extracted_text(self, extracted_text):
        is_fix_needed = len(extracted_text) % len(self.ocr_headers)
        if is_fix_needed == 0:
            return extracted_text

        header_size = len(self.ocr_headers)
        max_size = (len(extracted_text) // header_size) * header_size
        for i in range(0, max_size, header_size):
            # 직위
            if not extracted_text[i + 1] in ["서클장", "서클원", "부서클장"]:
                if CircleMemberManager().get_by_nickname(extracted_text[i]) is None:
                    extracted_text[i] += extracted_text[i + 1]
                del extracted_text[i + 1]

            # 주간 공헌도
            if not extracted_text[i + 2].isdigit():
                extracted_text[i + 1] += extracted_text[i + 2]
                del extracted_text[i + 2]

            # # 누적 공헌도
            # if not extracted_text[i + 3].isdigit():
            #     pass

            # # 상태
            # if not extracted_text[i + 4] == "오늘" and not extracted_text[i + 4].endswith("전"):
            #     pass

            # 레벨
            if not extracted_text[i + 5].startswith("Lv"):
                extracted_text[i + 4] += extracted_text[i + 5]
                del extracted_text[i + 5]

        return extracted_text

    def calculate_missing_weekly_contrib(self, data):
        circle_member_manager = CircleMemberManager()
        join_date = circle_member_manager.get_by_nickname(data["nickname"], "join_date")

        if join_date:
            return self.calculate_weekly_contrib_goal(join_date) - int(data["weekly_contrib"])

    def calculate_weekly_contrib_goal(self, join_date):
        if join_date is None:
            return None

        # 현재 날짜와 시간
        today = datetime.now()

        # 5시 기준으로 요일을 변경
        if today.hour < 5:
            today -= timedelta(days=1)  # 5시 이전이면 전날로 설정

        # join_date를 datetime 객체로 변환
        join_date = datetime.strptime(join_date, '%Y-%m-%d')

        # 이번 주 가입은 구분
        weekday = today.weekday()
        if join_date > today - timedelta(days=weekday):  # 이번 주에 가입한 경우
            days_since_join = (today - join_date).days + 1
        else:
            monday = today - timedelta(days=weekday)
            days_since_join = (today - monday).days + 1

        return MAX_DAILY_CONTRIB * days_since_join

    def calculate_total_contrib_goal(self, join_date):
        if join_date is None:
            return None

        # 가입일을 기준으로 목표 누적 공헌도 계산
        today = datetime.now()

        # 5시 기준으로 요일을 변경
        if today.hour < 5:
            today -= timedelta(days=1)  # 5시 이전이면 전날로 설정

        join_date = datetime.strptime(join_date, '%Y-%m-%d')  # 가입일 문자열을 날짜 형식으로 변환
        days_joined = (today - join_date).days

        # 목표 누적 공헌도 = 90 * 가입일 수
        return MAX_DAILY_CONTRIB * days_joined

    def _cell_style_handler(self, cell, row_data):
        circle_member_manager = CircleMemberManager()
        # nickname_idx = next((idx for idx, column in enumerate(self.excel_columns) if column.key == "nickname"), None)
        # status_idx = next((idx for idx, column in enumerate(self.excel_columns) if column.key == "status"), None)
        member: CircleMember = circle_member_manager.get_by_nickname(row_data["nickname"])

        if member is None:
            cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")  # 배경색
            return

        config = ConfigManager()
        # contrib_idx = next(
        #     (idx for idx, column in enumerate(self.excel_columns) if column.key == "missing_weekly_contrib"), None)
        if row_data["missing_weekly_contrib"] is not None and row_data["missing_weekly_contrib"] >= config.get(
                ConfigKeys.CONTRIB_LIMIT):
            cell.fill = PatternFill(start_color="FFDFDF", end_color="FFDFDF", fill_type="solid")  # 배경색

    def get_dynamic_ratio(self, hwnd=None):
        hwnd = hwnd if hwnd else window.find_window(ConfigManager().get(ConfigKeys.WINDOW_TITLE))
        _, _, client_width, client_height, _, _ = window.get_client_area(hwnd=hwnd)

        # 윈도우 해상도를 기준으로 비율 계산
        aspect_ratio = client_width / client_height

        # 가로 비율 (일반 해상도 vs 울트라와이드)
        if aspect_ratio >= 21 / 9:  # 울트라와이드 (21:9 이상)
            left_ratio = 1 / 21.5
            right_ratio = left_ratio * 1.3
            top_ratio = 0.061  # 상단 비율
            bottom_ratio = top_ratio * 0.76
        else:  # 일반 해상도 (16:9, 16:10)
            left_ratio = 1 / 16
            right_ratio = left_ratio * 1
            top_ratio = 0.083  # 상단 비율
            bottom_ratio = top_ratio * 0.75

        profile_ratio = top_ratio * 0.77

        print(f"🔍 감지된 화면 비율: {aspect_ratio:.2f}")

        return aspect_ratio, left_ratio, right_ratio, top_ratio, bottom_ratio, profile_ratio

    def _on_before_export(self, hwnd=None):
        pass

    def _on_after_export(self, hwnd=None):
        pass

    def _on_before_extract(self, hwnd=None):
        pass

    def _on_after_extract(self, hwnd=None):
        pass
